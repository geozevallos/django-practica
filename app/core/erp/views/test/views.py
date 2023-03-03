from django.views.generic import TemplateView
from django.http import JsonResponse
from django.utils.decorators import method_decorator
from django.views.decorators.csrf import csrf_exempt

from django.http.response import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from core.erp.models import Category

from core.erp.forms import TestForm
from core.erp.models import Product


class TestView(TemplateView):
    template_name: str = 'tests.html'

    @method_decorator(csrf_exempt)
    def dispatch(self, request, *args, **kwargs):
        return super().dispatch(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title']='Select anidados | Django'
        context['form']=TestForm()
        return context

    def post(self, request, *args, **kwargs):
        data = {}
        try:
            action = request.POST['action']
            if action == 'autocomplete':
                data = []
                termino = request.POST['term']
                products = Product.objects.filter(name__icontains=termino)
                if len(products) > 0:
                    for i in products:
                        item = i.toJSON()
                        item['value'] = i.name
                        data.append(item)
            elif action == 'search_product_id':
                data = []
                print(request.POST['id'])
                for product in Product.objects.filter(cate_id=request.POST['id']):
                    data.append({'id': product.id, 'name': product.name})
            else:
                data['error'] = 'No ha ingresado a ninguna opción'
        except Exception as e:
            data['error'] = str(e)
        print(data)
        return JsonResponse(data, safe=False)


############### VIEW FOR EXCEL
class ReportePersonalizadoExcel(TemplateView):
    def get(self, request, *args, **kwargs):
        query = Category.objects.all()
        wb = Workbook()
        flag = True # Para varias hojas
        cont = 1
        controlador = 4 # Posicion para ingresar texto
        for element in query:
            print(element)
            if flag:
                ws = wb.active
                ws.title = 'Hoja' + str(cont)
                flag = False
            else:
                ws = wb.create_sheet('Hoja' + str(cont))
            
            # Estilos de titulo
            ws['B1'].alignment = Alignment(horizontal='center', vertical='center')
            ws['B1'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"), top=Side(border_style="thin"), bottom=Side(border_style="thin"))
            ws['B1'].fill = PatternFill(start_color='66FFCC', end_color='66FFCC', fill_type="solid")
            ws['B1'].font = Font(name='Tahoma', size=12, bold=True)
            ws['B1'] = 'REPORTE PERSONALIZADO EN EXCEL EN DJANGO'
            
            # Unir celdas
            ws.merge_cells('B1:E1')

            #Cambiar ancho y alto
            ws.column_dimensions['B'].width = 20
            ws.column_dimensions['C'].width = 20
            ws.column_dimensions['D'].width = 20
            ws.column_dimensions['E'].width = 20
            ws.row_dimensions[1].height = 25

            ws['B3'].alignment = Alignment(horizontal='center', vertical='center')
            ws['B3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"), top=Side(border_style="thin"), bottom=Side(border_style="thin"))
            ws['B3'].fill = PatternFill(start_color='66FFCC', end_color='66FFCC', fill_type="solid")
            ws['B3'].font = Font(name='Tahoma', size=10, bold=True)
            ws['B3'] = 'NOMBRE'

            ws['C3'].alignment = Alignment(horizontal='center', vertical='center')
            ws['C3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"), top=Side(border_style="thin"), bottom=Side(border_style="thin"))
            ws['C3'].fill = PatternFill(start_color='66FFCC', end_color='66FFCC', fill_type="solid")
            ws['C3'].font = Font(name='Tahoma', size=10, bold=True)
            ws['C3'] = 'DESRIPCIÓN'

            # Mostrar data
            ws.cell(row=controlador, column=2).border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"), top=Side(border_style="thin"), bottom=Side(border_style="thin"))
            ws.cell(row=controlador, column=2).font = Font(name='Tahoma', size=10, bold=True)
            ws.cell(row=controlador, column=2).value = element.name

            ws.cell(row=controlador, column=3).border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"), top=Side(border_style="thin"), bottom=Side(border_style="thin"))
            ws.cell(row=controlador, column=3).font = Font(name='Tahoma', size=10, bold=True)
            ws.cell(row=controlador, column=3).value = element.description

            cont += 1
        
        # Nombre archivo
        filename = "Reportepersonalizado.xlsx"

        # Definir tipo de respuesta
        response = HttpResponse(headers={'Content-Type': 'application/ms-excel',
                                         'Content-Disposition': 'attachment; filename={}'.format(filename),
                                         })
        wb.save(response)
        return response